import numpy as np
import openpyxl
import power_system


class ExcelPowerSystemBuilder:
    def __init__(self, filename, bus_data_worksheet_name, line_data_worksheet_name, generator_neutral_impedance):
        self._workbook = openpyxl.load_workbook(filename, read_only=True)
        self._bus_data_worksheet = self._workbook[bus_data_worksheet_name]
        self._line_data_worksheet = self._workbook[line_data_worksheet_name]
        self._generator_neutral_impedance = generator_neutral_impedance

    def build_buses(self):
        result = []
        for row in self._bus_data_worksheet.iter_rows(row_offset=1):
            bus_number = row[0].value
            if not bus_number:
                break

            voltage_magnitude = row[8].value
            voltage_angle = row[9].value
            voltage = voltage_magnitude * np.exp(1j * np.deg2rad(voltage_angle))

            # Y_load = S* / |V|^2
            p_load = row[1].value or 0
            q_load = row[2].value or 0
            s_load = p_load + 1j * q_load
            load_admittance = np.conjugate(s_load) / voltage_magnitude ** 2

            gen_z0 = 1j * row[6].value or 0j
            gen_z1 = 1j * row[4].value or 0j
            gen_z2 = 1j * row[5].value or 0j
            gen_zn = self._generator_neutral_impedance if row[7].value == 1 else np.inf

            result.append(power_system.Bus(bus_number, voltage, load_admittance, gen_z0, gen_z1, gen_z2, gen_zn))

        return result

    def build_lines(self):
        result = []
        for row in self._line_data_worksheet.iter_rows(row_offset=1):
            source_bus_number = row[0].value
            destination_bus_number = row[1].value
            if not source_bus_number or not destination_bus_number:
                break

            r_distributed = row[2].value or 0
            x_distributed = row[3].value or 0
            z_distributed = r_distributed + 1j * x_distributed
            y_shunt = 2j * row[4].value or 0j  # multiply by 2 because input gives B / 2

            result.append(power_system.Line(source_bus_number, destination_bus_number, z_distributed, y_shunt))

        return result

    def build_system(self):
        return power_system.PowerSystem(self.build_buses(), self.build_lines())
